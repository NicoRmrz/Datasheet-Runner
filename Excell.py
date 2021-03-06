from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, colors, PatternFill, Color, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter
from openpyxl.chart import BarChart, Series, Reference
import openpyxl
import time
import datetime
import os
from PyQt5.QtCore import Qt, QThread, pyqtSignal, pyqtSlot, QObject, QSize, QFileInfo, QRegularExpression

'''
Class: Excel_Report
    Class to handle all excel file generation functions
    
Parameters: 
    QObject - inherits QObject attributes
'''
class Excel_Report(QObject):
    suffix = "_Report"
    path = ""
    reportExport = ""

    '''
    Function: __init__
		  initializes when class is called and set some excel attribute variables
    '''
    def __init__(self):
        self.thick_border = Side(border_style = "thick")
        self.thin_border = Side(border_style= 'thin')
        self.THICK =  Border( top=self.thick_border,left=self.thick_border, right=self.thick_border,bottom=self.thick_border)
        self.THIN = Border(top = None, left = self.thin_border, right = self.thin_border,bottom = self.thin_border)
        self.FULLTHIN = Border(top = self.thin_border, left = self.thin_border, right = self.thin_border,bottom = self.thin_border)
        self.fontStyle = Font(size = "14")
        self.largeFontStyle = Font(size = "20",bold=True)
        self.boldFont = Font(size = "14",bold=True)
        self.fillColor = PatternFill(fgColor=Color('CDCDCD'), fill_type = "solid")
        self.centerAlignment = Alignment( horizontal='center', vertical='center', wrap_text=True)
        self.alignRight = Alignment( horizontal='right', vertical='center', wrap_text=True)

    '''
    Function: startExcelSheet
		  Create excel report sheet and set heading

    Parameters: 
	  	exportPath - report export network path
	  	name       - input JSON file name
	  	serialNum  - serial number of DUT
	  	Protocol   - Protocol name from JSON file

    Returns: 
	  	openFileName - current open excel file
    '''
    def startExcelSheet(self, exportPath, name, serialNum, Protocol):
        self.path = exportPath + name+ "/"
        print(serialNum)

        #Create test folder path in report folder
        if not os.path.exists(self.path):
          os.makedirs(self.path)

        fileName = serialNum + self.suffix
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = serialNum

        # Write protocol name on header
        report = self.ws.cell(row=1, column=1)
        report.fill =  self.fillColor
        report.font  =  self.fontStyle
        report.value = "Test Report: " 
        report.border = self.THICK 

        # Write protocol name on header
        self.ws.merge_cells("B1:C1")
        report = self.ws["B1"]
        report1 = self.ws["C1"]
        report.fill =  self.fillColor
        report.font  =  self.boldFont
        report.value = Protocol 
        report.border = self.THICK 
        report1.border = self.THICK 
        report.alignment = self.centerAlignment

        # Write serial number name on header
        serNum = self.ws.cell(row=2, column=1)
        serNum.fill =  self.fillColor
        serNum.font  =  self.fontStyle
        serNum.value = "Serial Number: "
        serNum.border = self.THICK 

        # Write serial number name on header
        self.ws.merge_cells("B2:C2")
        serNum = self.ws["B2"]
        serNum1 = self.ws["C2"]
        serNum.fill =  self.fillColor
        serNum.font  =  self.boldFont
        serNum.value = serialNum
        serNum.border = self.THICK 
        serNum1.border =  self.THICK 
        serNum.alignment = self.centerAlignment

        openFileName = self.path + fileName
        return openFileName

    '''
    Function: addSingleHeader
        Create excel sheet single section heading

    Parameters: 
	  	Hrow         - row to enter header
	  	Hcol         - col to enter header
	  	title        - section  header title
    '''
    def addSingleHeader(self, Hrow, Hcol, title):
        # Create header title
        titleCell =  self.ws.cell(row=Hrow, column=Hcol)
        titleCell.value = title
        titleCell.font  =  self.boldFont
        titleCell.fill =  self.fillColor
        titleCell.alignment = self.centerAlignment
        titleCell.border = self.FULLTHIN 

    '''
    Function: addHeaderRow
        Create excel sheet section heading with title and list of title headings

    Parameters: 
	  	Hrow         - row to enter header
	  	Hcol         - col to enter header
	  	title        - section  header title
	  	headerList   - items to put on header row

    Return:
        next row
    '''
    def addHeaderRow(self, Hrow, Hcol, title, headerList):
        # Create header title
        headerLen = len(headerList)

        if (Hcol == 1):
            self.ws.merge_cells(start_row=Hrow, start_column=Hcol, end_row=Hrow, end_column=Hcol + headerLen - 1)
        else:
            self.ws.merge_cells(start_row=Hrow, start_column=Hcol, end_row=Hrow, end_column=Hcol + headerLen - 1)

        titleCell =  self.ws.cell(row=Hrow, column=Hcol)
        titleCell.value = title
        titleCell.font  =  self.boldFont
        titleCell.fill =  self.fillColor
        titleCell.alignment = self.centerAlignment

        # Next Row
        Hrow += 1

        # write in header
        headCol = Hcol
        for i in headerList:
            self.ws.cell(row=Hrow, column=headCol).value = i

            # Border cell
            self.ws.cell(row=Hrow-1, column=headCol).border = self.THICK # add border to header title
            self.ws.cell(row=Hrow, column=headCol).border = self.THIN 
                                                        
            # Align cell
            self.ws.cell(row=Hrow, column=headCol).alignment = self.centerAlignment

            # color header
            self.ws.cell(row=Hrow, column=headCol).fill =  self.fillColor                                                            
            headCol += 1
        
        Hrow += 1
        return Hrow

    '''
    Function: writeExcelEntry
		  Writes to excel sheet

    Parameters: 
	  	entry - value to write to excel sheet
	  	row - excel row to write to 
	  	col - excel col to write to 
    '''
    def writeExcelEntry(self, entry, row, col):
        enterCell = self.ws.cell(row=row, column=col)
        enterCell.value = entry

        # Align cell
        self.ws.cell(row, col).alignment = self.centerAlignment

        # create border per cell
        enterCell.border = self.THIN

    '''
    Function: writeSignature
		  Writes to excel sheet tester name/ sign/ date

    Parameters: 
	  	row - excel row to write to 
    '''
    def writeSignature(self, row):

        # write tester name
        username = os.environ.get('USERNAME')
        self.ws.cell(row=row, column=1).value = "Tested By:"
        self.ws.cell(row=row, column=1).alignment = self.alignRight
        self.ws.cell(row=row, column=2).value = username
        self.ws.cell(row=row, column=2).alignment = self.centerAlignment

        # signature
        self.ws.cell(row=row, column=3).value = "Sign:"
        self.ws.cell(row=row, column=3).alignment = self.alignRight
        self.ws.cell(row=row, column=4).value = "________________________"

        # write current date 
        datestring = datetime.datetime.fromtimestamp(time.time()).strftime('%B %d, %Y')
        self.ws.cell(row=row, column=5).value = "Date:"
        self.ws.cell(row=row, column=5).alignment = self.alignRight
        self.ws.cell(row=row, column=6).value = datestring
        self.ws.cell(row=row, column=6).alignment = self.centerAlignment

    '''
    Function: colorCellFail
		  Color Fails test Cell Red

    Parameters: 
	  	row - excel row 
	  	col - excel col 
    '''
    def colorCellFail(self, row, col):
        colorCell = self.ws.cell(row=row, column=col)
        fillColor = PatternFill(fgColor=colors.RED, fill_type = "solid")
        colorCell.fill = fillColor

    '''
    Function: startDataAnalysis
        Create excel data analysis sheet and set heading

    Parameters: 
	  	reportPath - path to export data analysis when finished
    
    Returns: 
	  	DAfileName - name of data analysis sheet
    '''
    def startDataAnalysis(self, reportPath):
        self.reportExport = reportPath
        DAfileName =  self.reportExport + "/Data_Analysis_Results"

        baseFolder = QFileInfo(reportPath)

        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "Data Analysis"
        self.ws.merge_cells("A1:C1")

        # Write data analysis name on header
        DAHeader = self.ws.cell(row=1, column=1)
        DAHeader.fill =  self.fillColor
        DAHeader.font  =  self.largeFontStyle
        DAHeader.value = "Data Analysis: " + baseFolder.baseName() 
        DAHeader.border = self.THICK 
        self.ws.cell(row=1, column=2).border = self.THICK 
        self.ws.cell(row=1, column=3).border = self.THICK 

        return DAfileName

    '''
    Function: parseReport
        parse input report and retreive all data and return data is resultDict

    Parameters: 
	  	inputReport - input report to parse data5
    
    Returns: 
	  	resultDict - dictionary of results from report
    '''
    def parseReport(self, inputReport):
        resultDict ={}
        sectionList = []
        minList = []
        maxList = []
        unitList = []
        valueList = []
        resultList = []
        serNum = ""
        lastTestRow = 0
        firstRow = 0
        foundLastRow = False

        wb = load_workbook(filename = inputReport)
        wsl = wb.active
        serNum = wsl.cell(row=2, column=2).value # get serial number from sheet
        lastRow = wsl.max_row  # get last row of report

        resultDict["Serial Number"] = serNum

        # first find procedure header row
        for row in range(1, lastRow):
            if (wsl.cell(row=row, column=1).value == "Section"):

                # get first row
                firstRow = row + 1

        # find last row of test procedure section
        for row in range(firstRow, lastRow):
            
            # find first empty row (if signature is added on bottom)
            if (wsl.cell(row=row, column=1).value == None and foundLastRow == False):
                lastTestRow = row 
                foundLastRow = True
            # else last row is end of table
            elif (lastTestRow == 0):
                lastTestRow = lastRow + 1

        # iterate through each test sections and append to lists
        for row in range(firstRow, lastTestRow):
            sectionList.append(wsl.cell(row=row, column=1).value)
            minList.append(wsl.cell(row=row, column=2).value)
            maxList.append(wsl.cell(row=row, column=3).value)
            unitList.append(wsl.cell(row=row, column=4).value)
            valueList.append(wsl.cell(row=row, column=5).value)
            resultList.append(wsl.cell(row=row, column=6).value)
 
        # add each list to dict
        resultDict["Section"] = sectionList
        resultDict["Min"] = minList
        resultDict["Max"] = maxList
        resultDict["Unit"] = unitList
        resultDict["Value"] = valueList
        resultDict["Result"] = resultList

        return resultDict

    '''
    Function: getTimestamp
		  Gets current time for timestamp

    Returns: 
	  	timestamp_return - current timestamp
    '''
    def getTimestamp(self):
        timestamp_return = datetime.datetime.fromtimestamp(time.time()).strftime('%Y-%m-%d_%H_%M_%S')
        return timestamp_return

    '''
    Function: getTestDataList
        Iterate populated Data Analysis excel worksheet.
        Append all tests with values into a list and return value in list

    Parameters: 
	  	numTests    - number of total tests
	  	numReports  - total number of samples
        rowStart    - first row to start parsing data
        colStart    - first col to start parsing data

    Returns: 
	  	outputList - List of all values from each test from all samples
    '''
    def getTestDataList(self, numTests, numReports, rowStart, colStart):

        outputList = [] 

        #  set reguar expression value to 
        isHex = QRegularExpression("0x+[0-9a-f]", QRegularExpression.CaseInsensitiveOption) # hex number

        # append all tests with values
        for test in range(0, numTests):
           
            dataList = [] 
            # append data from all reports
            for report in range(0, numReports):

                value = self.ws.cell(row=rowStart + test, column= colStart + report).value
                if (value != 'N/A' and value != None):

                    # check base and convert to int
                    if (isHex.match(str(value)).hasMatch()): 
                        value = int(value, 16)    # convert hex val to int
                    else:
                        value = float(value)

                dataList.append(value)

            # append to test list
            outputList.append(dataList)
        return outputList

    '''
    Function: getDataColumn
        Gets list of column with data anlysis
        Return column data in list

    Parameters: 
	  	length      - length of column
        rowStart    - first row to start parsing data
        colStart    - first col to start parsing data

    Returns: 
	  	columnList - List of all values from each test from all samples
    '''
    def getDataColumn(self, length, rowStart, colStart):
        columnList = []

        for test in range(0, length):
            value = self.ws.cell(row=rowStart + test, column=colStart).value

            if (value != 'N/A'):
                columnList.append(value)
       
        return columnList

    '''
    Function: createBarGraph
		  Create bar graph of current data

    Parameters: 
	  	firstRow - first row of data
	  	length - length of data
        col - column of data to graph
    '''
    def createBarGraph(self, firstRow, length, col):
        lastCol =  self.ws.max_column
        maxRow = length + firstRow - 1

        # print("Min: " + str(firstRow))
        # print("Max: " + str(maxRow))
        # print("SD column: " + str(col))
        
        data = Reference(self.ws, min_col=2, min_row=firstRow - 1, max_row=maxRow + 1, max_col=4)
        cats = Reference(self.ws, min_col=col, min_row=firstRow, max_row=maxRow)
        # print(data)
        # print(cats)

        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = "Standard Deviation"
        chart1.y_axis.title = 'Test number'
        chart1.x_axis.title = 'Sample length (mm)'
        chart1.add_data(data, titles_from_data=True)

        chart1.set_categories(cats)
        chart1.shape = 11
        # chart1.shape = 4
        self.ws.add_chart(chart1, "A10")

    '''
    Function: SaveSheet
        To save excel sheet

    Parameters: 
	  	fileToSave - chosen file to save

    Returns: 
	  	Final_Report_Name - saved report name to display on UI
    '''
    def SaveSheet(self, fileToSave):
        # To Auto Fit column width
        dims = {}
        for row in self.ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
                     
        for col, value in dims.items():
            # print (str(col) + ", " + str(value))
            # self.ws.column_dimensions[col].width = value+3
            self.ws.column_dimensions[get_column_letter(col)].width = value+3

        #Timestamps the file
        # gettime = self.getTimestamp()
        # Final_Report_Name = fileToSave + gettime + '.xlsx'
        Final_Report_Name = fileToSave + '.xlsx'
        self.wb.save(Final_Report_Name)

        self.suffix = "_Report"

        return Final_Report_Name
